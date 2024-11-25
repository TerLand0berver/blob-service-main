"""Text extraction implementations for different file types"""

import io
import os
import logging
from typing import Optional, Dict, Any, List
import fitz  # PyMuPDF
import docx
import chardet
import openpyxl
import csv
import xlrd
from app.utils.exceptions import ProcessingError
from app.utils.text_utils import text_config

logger = logging.getLogger(__name__)

class BaseExtractor:
    """Base class for text extractors"""
    
    def __init__(self, content: bytes, filename: str):
        self.content = content
        self.filename = filename
        self.config = text_config
        
    async def extract(self) -> str:
        """Extract text from content"""
        raise NotImplementedError
        
    def _detect_encoding(self, sample: bytes) -> str:
        """Detect text encoding"""
        result = chardet.detect(sample)
        return result['encoding'] or self.config.general['fallback_encoding']

class PDFExtractor(BaseExtractor):
    """PDF file text extractor"""
    
    async def extract(self) -> str:
        try:
            with fitz.open(stream=self.content, filetype="pdf") as doc:
                # Validate page count
                self.config.validate_pdf_extraction(len(doc))
                
                # Extract text from each page
                text_parts = []
                for page in doc:
                    text = page.get_text()
                    text_parts.append(text)
                
                return '\n\n'.join(text_parts)
                
        except Exception as e:
            logger.error(f"Error extracting PDF content: {str(e)}")
            raise ProcessingError(f"Failed to extract PDF content: {str(e)}")

class SpreadsheetExtractor(BaseExtractor):
    """Extract text from spreadsheet files"""
    
    def _extract_xlsx(self) -> str:
        """Extract text from Excel file using openpyxl"""
        wb = openpyxl.load_workbook(io.BytesIO(self.content), read_only=True)
        text_parts = []
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_texts = []
            
            # Count rows and columns
            row_count = 0
            max_col_count = 0
            for row in ws.rows:
                row_count += 1
                col_count = len([cell for cell in row if cell.value is not None])
                max_col_count = max(max_col_count, col_count)
                
                if row_count == 1:  # First row
                    self.config.validate_spreadsheet_extraction(0, col_count, 'xlsx')
            
            # Validate total rows
            self.config.validate_spreadsheet_extraction(row_count, max_col_count, 'xlsx')
            
            # Extract text
            for row in ws.rows:
                row_text = []
                for cell in row:
                    if cell.hyperlink:
                        cell_text = f"[{cell.value}]({cell.hyperlink.target})"
                    else:
                        cell_text = str(cell.value) if cell.value is not None else ''
                    row_text.append(cell_text)
                if any(text.strip() for text in row_text):  # Skip empty rows
                    sheet_texts.append('\t'.join(row_text))
            
            if sheet_texts:
                text_parts.append(f"Sheet: {sheet_name}")
                text_parts.append('\n'.join(sheet_texts))
        
        wb.close()
        return '\n\n'.join(text_parts)
    
    def _extract_xls(self) -> str:
        """Extract text from legacy Excel file using xlrd"""
        wb = xlrd.open_workbook(file_contents=self.content)
        text_parts = []
        
        for sheet_idx in range(wb.nsheets):
            sheet = wb.sheet_by_index(sheet_idx)
            sheet_texts = []
            
            # Validate dimensions
            self.config.validate_spreadsheet_extraction(sheet.nrows, sheet.ncols, 'xls')
            
            # Extract text
            for row_idx in range(sheet.nrows):
                row_text = []
                for col_idx in range(sheet.ncols):
                    cell = sheet.cell(row_idx, col_idx)
                    cell_text = str(cell.value) if cell.value else ''
                    row_text.append(cell_text)
                if any(text.strip() for text in row_text):  # Skip empty rows
                    sheet_texts.append('\t'.join(row_text))
            
            if sheet_texts:
                text_parts.append(f"Sheet: {wb.sheet_names()[sheet_idx]}")
                text_parts.append('\n'.join(sheet_texts))
        
        return '\n\n'.join(text_parts)
    
    def _extract_csv(self) -> str:
        """Extract text from CSV file"""
        # Try different encodings
        encodings = ['utf-8', 'gbk', 'latin1']
        text = None
        
        for encoding in encodings:
            try:
                text = self.content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        
        if text is None:
            # If all encodings fail, use chardet to detect encoding
            detected = chardet.detect(self.content)
            if detected and detected['encoding']:
                try:
                    text = self.content.decode(detected['encoding'])
                except UnicodeDecodeError:
                    text = self.content.decode('utf-8', errors='ignore')
            else:
                text = self.content.decode('utf-8', errors='ignore')
        
        # Parse CSV
        reader = csv.reader(text.splitlines())
        rows = []
        row_count = 0
        max_col_count = 0
        
        for row in reader:
            row_count += 1
            col_count = len([col for col in row if col.strip()])
            max_col_count = max(max_col_count, col_count)
            
            if row_count == 1:  # First row
                self.config.validate_spreadsheet_extraction(0, col_count, 'csv')
            
            if any(col.strip() for col in row):  # Skip empty rows
                rows.append('\t'.join(row))
        
        # Validate total rows
        self.config.validate_spreadsheet_extraction(row_count, max_col_count, 'csv')
        
        return '\n'.join(rows)

    async def extract(self) -> str:
        """Extract text from spreadsheet file"""
        try:
            ext = os.path.splitext(self.filename)[1].lower()
            
            if ext == '.xlsx':
                return self._extract_xlsx()
            elif ext == '.xls':
                return self._extract_xls()
            elif ext == '.csv':
                return self._extract_csv()
            else:
                raise ProcessingError(f"Unsupported spreadsheet format: {ext}")
            
        except Exception as e:
            logger.error(f"Error extracting spreadsheet content: {str(e)}")
            raise ProcessingError(f"Failed to extract spreadsheet content: {str(e)}")

class DocumentExtractor(BaseExtractor):
    """Document file text extractor"""
    
    async def extract(self) -> str:
        try:
            # Check file size
            size_mb = len(self.content) / (1024 * 1024)
            ext = os.path.splitext(self.filename)[1].lower()
            self.config.validate_document_extraction(size_mb, ext)
            
            if ext in ['.docx', '.doc']:
                doc = docx.Document(io.BytesIO(self.content))
                return '\n'.join(para.text for para in doc.paragraphs)
                
            elif ext in ['.txt', '.rtf']:
                encoding = self._detect_encoding(self.content[:4096])
                return self.content.decode(encoding)
            
            raise ProcessingError(f"Unsupported document format: {ext}")
            
        except Exception as e:
            logger.error(f"Error extracting document content: {str(e)}")
            raise ProcessingError(f"Failed to extract document content: {str(e)}")

class CodeExtractor(BaseExtractor):
    """Source code file text extractor"""
    
    async def extract(self) -> str:
        try:
            # Check file size
            size_mb = len(self.content) / (1024 * 1024)
            ext = os.path.splitext(self.filename)[1].lower()
            
            if ext[1:] not in self.config.code_config['supported_formats']:
                raise ProcessingError(f"Unsupported code format: {ext}")
                
            if size_mb > self.config.code_config['max_size_mb']:
                raise ProcessingError(
                    f"File exceeds maximum size of {self.config.code_config['max_size_mb']}MB"
                )
            
            # Detect encoding and extract text
            encoding = self._detect_encoding(self.content[:4096])
            return self.content.decode(encoding)
            
        except Exception as e:
            logger.error(f"Error extracting code content: {str(e)}")
            raise ProcessingError(f"Failed to extract code content: {str(e)}")

# Factory function to create appropriate extractor
def create_extractor(content: bytes, filename: str) -> BaseExtractor:
    """Create appropriate extractor based on file extension"""
    ext = os.path.splitext(filename)[1].lower()
    
    if ext in ['.pdf']:
        return PDFExtractor(content, filename)
    elif ext in ['.xlsx', '.xls', '.csv', '.tsv']:
        return SpreadsheetExtractor(content, filename)
    elif ext in ['.doc', '.docx', '.txt', '.rtf']:
        return DocumentExtractor(content, filename)
    elif ext[1:] in text_config.code_config['supported_formats']:
        return CodeExtractor(content, filename)
    else:
        raise ProcessingError(f"Unsupported file format: {ext}")

async def extract_text(content: bytes, filename: str) -> str:
    """Extract text from file content"""
    try:
        extractor = create_extractor(content, filename)
        return await extractor.extract()
    except Exception as e:
        logger.error(f"Error in text extraction: {str(e)}")
        raise ProcessingError(f"Text extraction failed: {str(e)}")
